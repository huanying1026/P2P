import os
import time
import pymysql
import logging
import smtplib
from configparser import ConfigParser
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# 当前文件路径
cur_path=os.path.dirname(os.path.realpath(__file__))
##写入日志
#------------------------------
#------------------------------
#类名：InsertLog_P
#目的：写入日志
#参数：无
#返回值：无
#创建时间：2019/05/04
#创建者:hjq
#修改者：
#修改时间：
#修改原因：
#------------------------------
#------------------------------
#log_path日志存放路径
log_path = os.path.realpath(os.path.join(os.path.dirname(cur_path),'log'))
class InsertLog_P():
    def __init__(self):
        self.logname= os.path.join(log_path,'%s.log'%time.strftime('%y-%m-%d'))
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        self.formatter =  logging.Formatter('[%(asctime)s-%(funcName)s line:%(lineno)3d-%(levelname)s:%(message)s]')

    def __console(self,level,message):
        #创建一个FileHandler,写入日志
        fh = logging.FileHandler(self.logname,'a',encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(self.formatter)
        self.logger.addHandler(fh)

        #创建一个StreamHandler，写入控制台显示
        sh=logging.StreamHandler()
        sh.setLevel(logging.DEBUG)
        sh.setFormatter(self.formatter)
        self.logger.addHandler(sh)

        if level=='info':
            self.logger.info(message)
        elif level=='debug':
            self.logger.debug(message)
        elif level=='warning':
            self.logger.warning(message)
        elif level=='error':
            self.logger.error(message)
        self.logger.removeHandler(fh)
        self.logger.removeHandler(sh)
        fh.close()

    def debug(self,message):
        self.__console('debug',message)

    def info(self,message):
        self.__console('info',message)

    def warning(self,message):
        self.__console('warning',message)

    def error(self,message):
        self.__console('error',message)


##操作数据库
#------------------------------
#------------------------------
#类名：OperateDataBase_P
#目的：操作数据库
#参数：无
#返回值：无
#创建时间：2019/05/04
#创建者:hjq
#修改者：
#修改时间：
#修改原因：
#------------------------------
#------------------------------
class OperateDataBase_P():
    def __init__(self,host,port,user,password,db,sql):
        self.host =host
        self.port=port
        self.user=user
        self.password=password
        self.db=db
        self.sql=sql

    def read_ini(self,filepath,section):
        c = ConfigParser()
        lis={}
        c.read(filepath,encoding='utf8')
        rets = c.options(section)
        for item in rets:
            value = c.get(section,item)
            lis[item] = value
        return lis

    def delete_data(self):
        try:
            conn = pymysql.connect(host=self.host,port=self.port,user=self.user,password=self.password,db=self.db)
            curs = conn.cursor()
            count = curs.execute(self.sql)
            conn.commit()
            curs.close()
            conn.close()
            return count
        except BaseException as msg:
            log = InsertLog_P()
            log.error(msg)

##获取不需要执行的模块(针对excel)
#------------------------------
#------------------------------
#类名：GetSkipScripts_P
#目的：获取不需要执行的模块
#参数：无
#返回值：无
#创建时间：2019/05/04
#创建者:hjq
#修改者：
#修改时间：
#修改原因：
#------------------------------
#------------------------------
class GetSkipScripts_P():
    def get_skip_scripts(self,filepath):
        try:
            m=[]
            wb = load_workbook(filepath)
            ws = wb['ScriptPath']
            rowcount= ws.max_row
            for i in range(2,rowcount+1):
                cellvalue = ws.cell(row=i,column=3).value
                if cellvalue=='False':
                    modelname =ws.cell(row=i,column=2).value
                    m.append(modelname)
            wb.close()
            return m
        except BaseException as msg:
            log = InsertLog_P()
            log.error(msg)

##获取不需要执行的用例(针对excel)
#------------------------------
#------------------------------
#类名：GetSkipTestCases_P
#目的：获取不需要执行的模块
#参数：无
#返回值：无
#创建时间：2019/05/04
#创建者:hjq
#修改者：
#修改时间：
#修改原因：
#------------------------------
#------------------------------
class GetSkipTestCases_P():
    def get_skip_testcases(self,FilePath):
        try:
            t = []
            wb = load_workbook(FilePath)
            sheels = wb.sheetnames
            # print sheels
            for i in sheels:
                ws = wb[i]
                rowcount = ws.max_row
                for j in range(2, rowcount + 1):
                    cellvalue = ws.cell(row=j, column=7).value
                    if cellvalue == 'False':
                        testcasename = ws.cell(row=j, column=1).value
                        t.append(testcasename)
            wb.close()
            return t
        except BaseException as msg:
            log = InsertLog_P()
            log.error(msg)

FD = "./reports"

def GetNewReport(FileDir=FD):
    # 打印目录所在所有文件名（列表对象）
    l = os.listdir(FileDir)
    l.sort(key=lambda fn: os.path.getmtime(FileDir + "\\" + fn))
    f = os.path.join(FileDir, l[-1])
    return f

class SendEmail_P():
    def __init__(self,sender,psw,receiver,smtpserver,report_file,port):
        self.sender=sender
        self.psw=psw
        self.receiver = receiver
        self.smtpserver=smtpserver
        self.report_file=report_file
        self.port=port

    def sendEmail(self):
        with open(self.report_file,'rb') as f:
            mail_body = f.read()
        #定义邮件内容
        msg = MIMEMultipart()
        body = MIMEText(mail_body,_subtype='html',_charset='utf-8')
        msg['Subject'] = '自动化测试报告'
        msg['from'] = self.sender
        msg['to'] = self.psw
        msg.attach(body)
        #添加附件
        att=MIMEText(open(self.report_file,'rb').read(),'base64','utf-8')
        att["Content-Type"] = "application/octet-stream"
        att["Content-Disposition"] = 'attachment; filename= "report.html"'
        msg.attach(att)
        try:
            smtp = smtplib.SMTP_SSL(self.smtpserver, self.port)
        except:
            smtp = smtplib.SMTP()
            smtp.connect(self.smtpserver, self.port)
        #用户名，授权密码
        smtp.login(self.sender,self.psw)
        smtp.sendmail(self.sender, self.receiver, msg.as_string())
        smtp.quit()
        print('邮件发送成功！')

if __name__ == '__main__':
    pass